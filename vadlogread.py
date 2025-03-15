import os
import re
from openpyxl import Workbook, load_workbook

# 检查并创建或读取Excel文件
def check_or_create_excel(file_name):
    if os.path.exists(file_name):
        workbook = load_workbook(file_name)
        sheet = workbook.active
        if sheet.title != "Sheet1":
            sheet.title = "Sheet1"
        # 检查表头是否正确
        if sheet.max_row < 1 or sheet.cell(row=1, column=1).value != "begin":
            sheet.append(["begin", "end", "r_text"])  # 创建表头
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.append(["begin", "end", "r_text"])  # 创建表头
    return workbook, sheet

# 读取日志文件并提取信息
def extract_log_data(log_file_path):
    with open(log_file_path, "r", encoding="utf-8") as file:
        lines = file.readlines()
    
    data = []
    begin_time = None
    r_text = ""  # 初始化r_text为空
    for i, line in enumerate(lines):
        if "onResult FIND_BOS." in line:
            begin_time = re.match(r"(\d{2}:\d{2}:\d{2})", line)
            if begin_time:
                begin_time = begin_time.group(1)
                print(f"Found begin time: {begin_time}")
            else:
                print(f"Failed to extract begin time from line: {line.strip()}")
        elif "onResult FINISH" in line and begin_time:
            end_time = re.match(r"(\d{2}:\d{2}:\d{2})", line)
            if end_time:
                end_time = end_time.group(1)
                print(f"Found end time: {end_time}")
                # 查找FinalrequestText
                for next_line in lines[i + 1:]:
                    if "onResult FIND BOS." in next_line:
                        break  # 如果遇到下一个onResult FIND BOS，停止查找
                    if "FinalrequestText" in next_line:
                        r_text_match = re.search(r"FinalrequestText\s*=\s*(.*)", next_line)
                        if r_text_match:
                            r_text = r_text_match.group(1).strip()
                            print(f"Found FinalrequestText: {r_text}")
                            break
                data.append([begin_time, end_time, r_text])  # 添加数据到列表
                begin_time = None
                r_text = ""  # 重置r_text
            else:
                print(f"Failed to extract end time from line: {line.strip()}")
    return data

# 主程序
def main():
    excel_file_name = "resulttime.xlsx"
    log_folder_name = "log"
    
    # 检查并创建或读取Excel文件
    workbook, sheet = check_or_create_excel(excel_file_name)
    
    # 检查是否存在log文件夹
    if not os.path.exists(log_folder_name):
        print(f"Error: The folder '{log_folder_name}' does not exist.")
        return
    
    # 遍历log文件夹中的所有.log文件
    log_files = [f for f in os.listdir(log_folder_name) if f.endswith(".log")]
    if not log_files:
        print(f"No .log files found in the folder '{log_folder_name}'.")
        return
    
    for log_file in log_files:
        log_file_path = os.path.join(log_folder_name, log_file)
        print(f"Processing log file: {log_file}")
        data = extract_log_data(log_file_path)
        if data:
            for row in data:
                sheet.append(row)
        else:
            print(f"No data extracted from {log_file}.")
    
    # 保存Excel文件
    workbook.save(excel_file_name)
    print(f"Data has been successfully written to {excel_file_name}")

if __name__ == "__main__":
    main()
